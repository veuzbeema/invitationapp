from django.shortcuts import render
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
from .models import Invitation
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import json
from .forms import InvitationForm
from django import forms

# views.py (complete functional views)
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
from .models import Invitation, InvitationCSVUpload
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.conf import settings
import json
from datetime import date
import csv
from io import StringIO
from events.models import Event, TicketClass
import uuid
from datetime import datetime
import re
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from django.core.exceptions import ValidationError
import re

from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum
from .models import Invitation  , RegisteredUser, InvitationCSVUpload
from django.conf import settings
from django.core.mail import send_mail
from django.utils import timezone
from datetime import datetime, date
import json
import re
import csv
from events.models import ExhibitorTicketAllocation
import logging

from invitations.tasks import send_invitation_email,send_bulk_invitations

logger = logging.getLogger(__name__)

# Create your views here.
class FileUploadForm(forms.Form):
    file = forms.FileField()

@csrf_exempt
@require_http_methods(["POST", "GET"])
def file_upload_view(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['file']
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            file_url = fs.url(filename)
            return render(request, 'invitations/file_upload_success.html', {'file_url': file_url})
    else:
        form = FileUploadForm()
    return render(request, 'invitations/file_upload_form.html', {'form': form})

# def invitation_list_view(request):
#     invitations = Invitation.objects.all()
#     return render(request, 'invitations/invitation_list.html', {'invitations': invitations})

def invitation_view(request, pk):
    invitation = get_object_or_404(Invitation, pk=pk)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        data = {
            'event': invitation.event.name if invitation.event else '',
            'title_or_name': invitation.title_or_name or '',
            'email': invitation.email or '',
            'invite_type': invitation.get_invite_type_display(),
            'status': invitation.get_status_display(),
            'expiry_date': invitation.expiry_date.strftime('%Y-%m-%d %H:%M:%S'),
            'link_limit': invitation.link_limit,
            'link_count': invitation.link_count,
            'registered_count': invitation.registered_count,
            'company_name': invitation.company_name or '',
            'personal_message': invitation.personal_message or '',
        }
        return JsonResponse(data)
    return JsonResponse({'error': 'Invalid request'}, status=400)


@require_http_methods(["GET"])
def invitation_get(request, pk):
    """
    Get invitation data for editing
    """
    invitation = get_object_or_404(Invitation, pk=pk)
    
    # Format expiry_date for date input (YYYY-MM-DD)
    expiry_date_formatted = invitation.expiry_date.strftime('%Y-%m-%d') if invitation.expiry_date else ''
    
    data = {
        'id': invitation.id,
        'title_or_name': invitation.title_or_name,
        'email': invitation.email,
        'ticket_type': invitation.ticket_class.ticket_type if invitation.ticket_class else '',
        'company_name': invitation.company_name,
        'personal_message': invitation.personal_message,
        'expiry_date': expiry_date_formatted,
        'link_limit': invitation.link_limit,
        'status': invitation.status,
        'invite_type': invitation.invite_type
    }
    
    return JsonResponse({'success': True, 'data': data})


@csrf_exempt  
@require_http_methods(["POST"])
def invitation_edit(request, pk):
    """
    Update invitation data
    """
    try:
        invitation = get_object_or_404(Invitation, pk=pk)
        data = json.loads(request.body)
        
        # Update fields
        invitation.title_or_name = data.get('title_or_name', invitation.title_or_name)
        invitation.company_name = data.get('company_name', invitation.company_name)
        invitation.personal_message = data.get('personal_message', invitation.personal_message)
        invitation.link_limit = int(data.get('link_limit', invitation.link_limit))
        invitation.status = data.get('status', invitation.status)
        
        # Handle expiry date
        expiry_date_str = data.get('expiry_date')
        if expiry_date_str:
            from datetime import datetime
            invitation.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d')
        
        # Handle ticket type
        ticket_type = data.get('ticket_type')
        if ticket_type:
            try:
                ticket_class = TicketClass.objects.get(event=invitation.event, ticket_type=ticket_type)
                invitation.ticket_class = ticket_class
            except TicketClass.DoesNotExist:
                return JsonResponse({'success': False, 'error': f'Invalid ticket type: {ticket_type}'}, status=400)
        
        invitation.save()
        
        return JsonResponse({'success': True, 'message': 'Invitation updated successfully'})
    
    except ValueError as e:
        return JsonResponse({'success': False, 'error': 'Invalid data format'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Delete invitation
def invitation_delete(request, pk):
    invitation = get_object_or_404(Invitation, pk=pk)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        invitation.delete()
        return JsonResponse({'message': 'Invitation deleted successfully'})
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
@require_http_methods(["GET"])
def invitation_list_view(request):
    # ✅ Handle AJAX (table filtering + searching)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            invitations = Invitation.objects.filter(
                event__created_by=request.user
            ).order_by('-created_at')

            # 🔸 Extract filters/search
            keyword = request.GET.get('keyword', '').strip()
            status = request.GET.get('status', '').strip()   # all / active / expired / pending
            invite_type = request.GET.get('type', '').strip()
            expiry = request.GET.get('expiry', '').strip()

            # 🔸 Keyword search
            if keyword:
                invitations = invitations.filter(
                    Q(title_or_name__icontains=keyword) |
                    Q(email__icontains=keyword) |
                    Q(invitation_key__icontains=keyword) |
                    Q(company_name__icontains=keyword)
                )

            # 🔸 Status filter
            if status and status != 'all':
                now = datetime.now()
                if status == 'active':
                    invitations = invitations.filter(
                        status='active',
                        expiry_date__gte=now
                    )
                elif status == 'expired':
                    invitations = invitations.filter(
                        expiry_date__lt=now
                    )
             
            # 🔸 Type filter (map UI → DB)
            if invite_type:
                type_map = {
                    'link': 'private_link',
                    'personal': 'personalized'
                }
                mapped_type = type_map.get(invite_type, invite_type)
                invitations = invitations.filter(invite_type=mapped_type)

            # 🔸 Expiry date filter
            if expiry:
                invitations = invitations.filter(expiry_date__date=expiry)

            # ✅ Build JSON data for table
            invitations_data = []
            for inv in invitations:
                invitations_data.append({
                    'name': inv.title_or_name or 'N/A',
                    'email': inv.email or 'N/A',
                    'type': inv.get_invite_type_display() or inv.invite_type,
                    'expiry': inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
                    'limit': inv.link_limit or 0,
                    'registered': inv.registered_count or 0,
                    'status': inv.status.capitalize(),
                    'key': inv.invitation_key or ''
                })

            logger.debug(f"Returning {len(invitations_data)} invitations after filtering")
            return JsonResponse({'data': invitations_data}, status=200)

        except Exception as e:
            logger.error(f"Error in invitation_list AJAX: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)

    # ✅ Normal page load
    try:
        # Ticket stats
        total_tickets = ExhibitorTicketAllocation.objects.filter(
            exhibitor__event__created_by=request.user
        ).aggregate(total=Sum('quantity'))['total'] or 0

        used_tickets = ExhibitorTicketAllocation.objects.filter(
            exhibitor__event__created_by=request.user, is_used=True
        ).aggregate(total=Sum('quantity'))['total'] or 0

        available_tickets = total_tickets - used_tickets

        # Ticket types (for stat cards)
        ticket_types = []
        ticket_classes = TicketClass.objects.filter(event__created_by=request.user)
        for ticket_class in ticket_classes:
            total = ExhibitorTicketAllocation.objects.filter(
                ticket_class=ticket_class,
                exhibitor__event__created_by=request.user
            ).aggregate(total=Sum('quantity'))['total'] or 0
            used = ExhibitorTicketAllocation.objects.filter(
                ticket_class=ticket_class,
                exhibitor__event__created_by=request.user,
                is_used=True
            ).aggregate(total=Sum('quantity'))['total'] or 0
            ticket_types.append({
                'name': ticket_class.ticket_type,
                'used': used,
                'total': total,
                'color_class': {
                    'visitor': 'info',
                    'vip': 'success',
                    'gold': 'warning',
                    'platinum': 'primary',
                    'exhibitor': 'secondary'
                }.get(ticket_class.ticket_type.lower(), 'primary')
            })

        # 🔸 Registered visitors count (sum of registered_count)
        registered_visitor = Invitation.objects.filter(
            event__created_by=request.user
        ).aggregate(total=Sum('registered_count'))['total'] or 0

        # Other stats
        generated_invitations = Invitation.objects.filter(event__created_by=request.user).count()

        # ✅ Context for template
        context = {
            'today': date.today(),
            'total_tickets': total_tickets,
            'used_tickets': used_tickets,
            'available_tickets': available_tickets,
            'allocated_invitations': total_tickets,
            'generated_invitations': generated_invitations,
            'registered_visitor': registered_visitor,   # 🔸 matches template variable
            'ticket_types': ticket_types,
        }
        return render(request, 'invitations/invitation_list.html', context)

    except Exception as e:
        logger.error(f"Error rendering invitation_list: {str(e)}")
        return render(request, 'invitations/invitation_list.html', {'error': str(e)})



def invitation_create_view(request):
    if request.method == 'POST':
        form = InvitationForm(request.POST, request.FILES)
        if form.is_valid():
            invitation = form.save()
            return render(request, 'invitations/invitation_detail.html', {'invitation': invitation})
    else:
        form = InvitationForm()
    return render(request, 'invitations/invitation_form.html', {'form': form})

def invitation_update_view(request, invitation_id):
    try:
        invitation = Invitation.objects.get(id=invitation_id)
    except Invitation.DoesNotExist:
        return render(request, 'invitations/not_found.html', status=404)
    if request.method == 'POST':
        form = InvitationForm(request.POST, request.FILES, instance=invitation)
        if form.is_valid():
            invitation = form.save()
            return render(request, 'invitations/invitation_detail.html', {'invitation': invitation})
    else:
        form = InvitationForm(instance=invitation)
    return render(request, 'invitations/invitation_form.html', {'form': form, 'invitation': invitation})

def invitation_delete_view(request, invitation_id):
    try:
        invitation = Invitation.objects.get(id=invitation_id)
    except Invitation.DoesNotExist:
        return render(request, 'invitations/not_found.html', status=404)
    if request.method == 'POST':
        invitation.delete()
        return render(request, 'invitations/invitation_deleted.html')
    return render(request, 'invitations/invitation_confirm_delete.html', {'invitation': invitation})





def validate_email(email):
    pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    return re.match(pattern, email) is not None



@csrf_exempt
@require_http_methods(["POST"])
def send_private_invitation(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)

        print('-------------data--------', data)
        link_title = data.get('linkTitle', '')
        link_limit = int(data.get('linkLimit'))
        ticket_type = data.get('TicketType')
        link_count = int(data.get('linkCount'))
        expire_date_str = data.get('expireDate', '')
        
        if not all([link_title, link_limit, ticket_type, expire_date_str, link_count]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
        
        if link_count > 1000 or link_limit > 100:
            return JsonResponse({'success': False, 'error': 'Link count or link limit exceeds maximum allowed'}, status=400)

        # Parse expire date
        expire_date = datetime.strptime(expire_date_str, '%Y-%m-%d').replace(tzinfo=None)
        
        # Get event and ticket class
        event = Event.objects.first()
        if not event:
            return JsonResponse({'success': False, 'error': 'No event found'}, status=500)
        
        try:
            ticket_class = TicketClass.objects.get(event=event, ticket_type=ticket_type)
        except TicketClass.DoesNotExist:
            return JsonResponse({'success': False, 'error': f'Invalid ticket type: {ticket_type}'}, status=400)
        
        base_url = request.build_absolute_uri('/').rstrip('/')
        
        # Generate multiple invitation links
        created_invitations = []
        generated_links = []
        
        for i in range(link_count):
            # Generate unique invitation key for each link
            invitation_key = str(uuid.uuid4()).replace('-', '')[:12].upper()
            
            # Create invitation
            invitation = Invitation.objects.create(
                event=event,
                title_or_name=link_title,
                invite_type='private_link',
                expiry_date=expire_date,
                link_limit=link_limit,  # How many times THIS specific link can be used
                link_count=1,  # This is a single link (part of a batch)
                registered_count=0,  # Start at 0, will increment as people register
                invitation_key=invitation_key,
                status='active',
                ticket_class=ticket_class,
            )
            print(invitation.status, "===========status================")
            
            created_invitations.append(invitation.id)
            generated_links.append({
                'id': invitation.id,
                'key': invitation_key,
                'url': f'{base_url}/register/{invitation_key}/'
            })
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {link_count} invitation link(s)',
            'invitation_ids': created_invitations,
            'links': generated_links,
            'total_count': link_count
        })
            
    except ValueError as e:
        return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def send_personalized_invitation(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    """
    try:
        data = json.loads(request.body)
        guest_name = data.get('guestName')
        guest_email = data.get('guestEmail')
        ticket_type = data.get('ticketType')
        company_name = data.get('companyName', '')
        personal_message = data.get('personalMessage', '')
        expire_date_str = data.get('personalExpireDate')


        print('-------------data--------', data)
        
        if not all([guest_name, guest_email, ticket_type, expire_date_str]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

        # Validate guest_name
        if len(guest_name) < 2:
            return JsonResponse({'success': False, 'error': 'Guest name must be at least 2 characters long'}, status=400)
        if len(guest_name) > 255:
            return JsonResponse({'success': False, 'error': 'Guest name cannot exceed 255 characters'}, status=400)
        if not re.match(r"^[A-Za-z\s\-\']+$", guest_name):
            return JsonResponse({
                'success': False,
                'error': 'Guest name can only contain letters, spaces, hyphens, or apostrophes'
            }, status=400)

        
        if not validate_email(guest_email):
            return JsonResponse({'success': False, 'error': 'Invalid email'}, status=400)
        
        # Parse expire date
        expire_date = datetime.strptime(expire_date_str, '%Y-%m-%d').replace(tzinfo=None)

        print('-------------expire_date--------', expire_date, datetime.now())
        if expire_date < datetime.now():
            return JsonResponse({'success': False, 'error': 'Expire date must be in the future'}, status=400)
        
        
        # Get event and ticket class (assuming first event for demo; adjust to user's event)
        event = Event.objects.first()
        if not event:
            return JsonResponse({'success': False, 'error': 'No event found'}, status=500)
        
        try:
            ticket_class = TicketClass.objects.get(event=event,ticket_type=ticket_type)
        except TicketClass.DoesNotExist:
            return JsonResponse({'success': False, 'error': f'Invalid ticket type: {ticket_type}'}, status=400)
        
        # Generate invitation key
        invitation_key = str(uuid.uuid4()).replace('-', '')[:12].upper()
        
        # Create invitation
        invitation = Invitation.objects.create(
            event=event,
            title_or_name=guest_name,
            email=guest_email,
            invite_type='personalized',
            expiry_date=expire_date,
            link_limit=1,
            invitation_key=invitation_key,
            status='active',
            ticket_class=ticket_class,
            company_name=company_name,
            personal_message=personal_message,
        )
        print(invitation.status, "===========status================")
        
        # Send email
        subject = f'Invitation to GITEX GLOBAL 2025 - {ticket_type.upper()} Pass'
        message_body = f"""
        Dear {guest_name},

        {personal_message or 'We are pleased to invite you to GITEX GLOBAL 2025.'}

        Please register using this link: localhost:8000/invite/{invitation_key}

        This invitation expires on: {expire_date_str}
        Company: {company_name or 'N/A'}

        Best regards,
        GITEX Team
                
                
        """
        sent = send_mail(
            subject,
            message_body,
            settings.DEFAULT_FROM_EMAIL,
            [guest_email],
            fail_silently=False,
        )
        # sent = True
        
        if sent:
            return JsonResponse({
                'success': True,
                'message': f'Invitation sent to {guest_email}',
                'invitation_id': invitation.id,
                'link': f'https://gitex.com/invite/{invitation_key}'
            })
        else:
            return JsonResponse({'success': False, 'error': 'Failed to send email'}, status=500)
            
    except ValueError as e:
        return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def send_bulk_personalized_invitation(request):
    """
    Handle bulk personalized invitations via AJAX POST with file upload.
    """
    try:
        # Get form data
        csv_file = request.FILES.get('bulkCsvFile')
        default_message = request.POST.get('bulkPersonalMessage', '')
        expire_date_str = request.POST.get('bulkExpireDate')
        is_large_file = request.POST.get('isLargeFile', 'false').lower() == 'true'
        
        if not csv_file or not expire_date_str:
            return JsonResponse({'success': False, 'error': 'Missing CSV file or expire date'}, status=400)
        
        # Parse expire date
        expire_date = datetime.strptime(expire_date_str, '%Y-%m-%d').replace(tzinfo=None)

        if expire_date < datetime.now():
            return JsonResponse({'success': False, 'error': 'Expire date must be in the future'}, status=400)
        
        # Get event
        event = Event.objects.first()
        if not event:
            return JsonResponse({'success': False, 'error': 'No event found'}, status=500)
        
        # Initialize variables for duplicate checking
        emails_to_process = []
        email_counts = {}
        duplicate_emails = []
        errors = []

        # Read and parse CSV for duplicate email check
        csv_file.seek(0)  # Reset file pointer
        csv_data = csv_file.read().decode('utf-8')
        csv_reader = csv.DictReader(StringIO(csv_data))

        # Validate headers
        expected_headers = ['Full Name', 'Email', 'Ticket Type', 'Company Name', 'Personal Message']
        if not all(header in csv_reader.fieldnames for header in expected_headers[:3]):
            return JsonResponse({'success': False, 'error': 'Invalid CSV headers. Expected: Full Name, Email, Ticket Type'}, status=400)

        # Collect emails for duplicate check
        for row in csv_reader:
            email = row.get('Email', '').strip().lower()
            if email and validate_email(email):
                emails_to_process.append(email)
                email_counts[email] = email_counts.get(email, 0) + 1

        # Identify duplicates within the CSV
        csv_duplicates = [email for email, count in email_counts.items() if count > 1]
        if csv_duplicates:
            errors.append(f'Warning: Duplicate emails found in CSV: {", ".join(csv_duplicates[:5])}{"..." if len(csv_duplicates) > 5 else ""}')
            duplicate_emails.extend(csv_duplicates)

        # Check existing emails in the database
        existing_invitations = Invitation.objects.filter(
            event=event,
            email__in=emails_to_process
        ).values_list('email', flat=True)
        existing_emails = list(existing_invitations)
        if existing_emails:
            errors.append(f'Warning: Emails already exist in database: {", ".join(existing_emails[:5])}{"..." if len(existing_emails) > 5 else ""}')
            duplicate_emails.extend(existing_emails)

        # Remove the redundant duplicate check that caused the error
        # The following block was incorrectly placed and caused the 'emails_to_process' error
        # It’s already handled above, so we don’t need it again

        valid_ticket_types = ['visitor', 'vip', 'gold', 'platinum', 'exhibitor']
        valid_rows = 0
        created_invitations = []

        if not is_large_file:
            # Reset reader for processing
            csv_file.seek(0)
            csv_data = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_data))
            
            for row_num, row in enumerate(csv_reader, start=2):
                name = row.get('Full Name', '').strip()
                email = row.get('Email', '').strip().lower()
                ticket_type = row.get('Ticket Type', '').strip().lower()
                company = row.get('Company Name', '').strip()
                personal_message = row.get('Personal Message', default_message).strip()

                if not (name and email and ticket_type):
                    errors.append(f'Row {row_num}: Missing required fields (Full Name, Email, Ticket Type)')
                    continue
                
                if not validate_email(email):
                    errors.append(f'Row {row_num}: Invalid email "{email}"')
                    continue
                
                if ticket_type not in valid_ticket_types:
                    errors.append(f'Row {row_num}: Invalid ticket type "{ticket_type}". Must be one of {valid_ticket_types}')
                    continue
                
                if email in duplicate_emails:
                    errors.append(f'Row {row_num}: Skipping duplicate email "{email}"')
                    continue
                
                try:
                    ticket_class = TicketClass.objects.get(event=event, ticket_type=ticket_type)
                except TicketClass.DoesNotExist:
                    errors.append(f'Row {row_num}: Ticket class "{ticket_type}" not found')
                    continue
                
                # Generate key
                invitation_key = str(uuid.uuid4()).replace('-', '')[:12].upper()
                
                # Create invitation
                invitation = Invitation.objects.create(
                    event=event,
                    title_or_name=name,
                    email=email,
                    invite_type='personalized',
                    expiry_date=expire_date,
                    link_limit=1,
                    link_count=1,
                    registered_count=0,
                    invitation_key=invitation_key,
                    status='active',
                    ticket_class=ticket_class,
                    company_name=company,
                    personal_message=personal_message,
                )
                
                created_invitations.append(invitation.id)
                valid_rows += 1
                
                send_invitation_email.delay(invitation.id)
            
            # Create CSV upload record AFTER processing
            csv_upload = InvitationCSVUpload.objects.create(
                event=event,
                file=csv_file,
                status='success' if not errors else ('partial' if valid_rows > 0 else 'failed'),
                processed=True,
                processed_at=datetime.now(),
                processed_count=valid_rows,
                failed_count=len(errors),
                error_message='\n'.join(errors[:100]) if errors else ''
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully processed {valid_rows} invitations. Errors: {len(errors)}',
                'valid_count': valid_rows,
                'error_count': len(errors),
                'errors': errors[:10],
                'csv_upload_id': csv_upload.id,
                'created_invitations': created_invitations
            })
        
        else:
            # For large files - create CSV record first, then process in background
            csv_upload = InvitationCSVUpload.objects.create(
                event=event,
                file=csv_file,
                status='processing',
                processed=False,
                processed_count=0,
                failed_count=0
            )
            
            send_bulk_invitations.delay(csv_upload.id, expire_date_str, default_message)
            
            return JsonResponse({
                'success': True,
                'message': 'Large file is being processed in the background.',
                'csv_upload_id': csv_upload.id,
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

from django.http import JsonResponse, FileResponse, HttpResponse
from .models import ExportJob

def exports_jobs(request):
    
    jobs = ExportJob.objects.all()
    return render(request, 'invitations/exports_jobs.html', {"jobs": jobs})


@login_required
def export_job_status(request, task_id):
    task = AsyncResult(task_id)
    export_job = ExportJob.objects.get(id=task.result.get('job_id')) if task.ready() and task.result else None
    
    response = {
        'task_id': task_id,
        'status': task.status,
        'job_status': export_job.status if export_job else 'pending',
        'progress': export_job.progress if export_job else 0,
        'download_url': f'/exports/download/{export_job.id}/' if export_job and export_job.status == 'completed' else None,
        'error_message': export_job.error_message if export_job and export_job.status == 'failed' else None
    }
    return JsonResponse(response)

@login_required
def export_job_download(request, job_id):
    try:
        export_job = ExportJob.objects.get(id=job_id)
        if export_job.status != 'completed' or not export_job.file:
            return HttpResponse('File not available', status=404)
        return FileResponse(export_job.file)
    except ExportJob.DoesNotExist:
        return HttpResponse('File not found', status=404)

@login_required
@require_http_methods(["GET"])
def invitation_list(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # AJAX request for DataTables
        try:
            # Fetch invitations for the logged-in user
            invitations = Invitation.objects.filter(event__created_by=request.user).order_by('-created_at')
            invitations_data = []

            for inv in invitations:
                invitations_data.append({
                    'id': inv.id,
                    'name': inv.title_or_name or 'N/A',
                    'email': inv.email or 'N/A',
                    'type': inv.invite_type,  # Use display for choices
                    'expiry': inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
                    'limit': inv.link_limit or 0,
                    'registered': inv.registered_count or 0,
                    
                    'status': inv.status.capitalize(),
                    'key': inv.invitation_key or ''
                })

            
            logger.debug(f"Returning {len(invitations_data)} invitations for DataTables")



            return JsonResponse({'data': invitations_data}, status=200)
        except Exception as e:
            logger.error(f"Error in invitation_list AJAX: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
        
    # allocated_invitations =  ExhibitorTicketAllocation.objects.all().aggregate(total=Sum('quantity'))['total'] or 0
    allocated_invitations = 20000
    generated_invitations = Invitation.objects.filter().count()
    

    context = {
        'total_tickets': 120,
        'used_tickets': 110,
        'available_tickets': 111,
        'allocated_invitations': allocated_invitations, 
        'generated_invitations': generated_invitations, 
        'remaining_invitations': allocated_invitations - generated_invitations,
        'registered_visitors': RegisteredUser.objects.filter().count(),
        'ticket_types': [],
        'invitation_base_url': 'http://localhost:8000/' if settings.DEBUG else settings.INVITATION_BASE_URL,
    }
    return render(request, 'invitations/invitation_list.html', context)
    # except Exception as e:
    #     logger.error(f"Error rendering invitation_list: {str(e)}")
    #     return render(request, 'invitations/invitation_list.html', {'error': str(e)})
    


@csrf_exempt
@require_http_methods(["POST"])
def edit_invitation(request):
    """
    Edit an existing invitation via AJAX POST.
    """
    try:
        data = json.loads(request.body)
        invitation_id = data.get('invitationId')
        title_or_name = data.get('guestName')
        ticket_type = data.get('ticketType')
        company_name = data.get('companyName')
        personal_message = data.get('personalMessage')
        expire_date_str = data.get('expireDate')
        link_limit = int(data.get('linkLimit', 1))
        status = data.get('status')

        if not all([invitation_id, title_or_name, ticket_type, expire_date_str, link_limit, status]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

        if link_limit > 100:
            return JsonResponse({'success': False, 'error': 'Link limit exceeds maximum (100)'}, status=400)

        try:
            invitation = Invitation.objects.get(id=invitation_id)
        except Invitation.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invitation not found'}, status=404)

        event = Event.objects.first()
        if not event:
            return JsonResponse({'success': False, 'error': 'No event found'}, status=500)

        try:
            ticket_class = TicketClass.objects.get(event=event, ticket_type=ticket_type)
        except TicketClass.DoesNotExist:
            return JsonResponse({'success': False, 'error': f'Invalid ticket type: {ticket_type}'}, status=400)

        try:
            expire_date = datetime.strptime(expire_date_str, '%Y-%m-%d').replace(tzinfo=None)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)

        # Update invitation fields
        invitation.title_or_name = title_or_name
        invitation.ticket_class = ticket_class
        invitation.company_name = company_name
        invitation.personal_message = personal_message
        invitation.expiry_date = expire_date
        invitation.link_limit = link_limit
        invitation.status = status
        invitation.save()

        return JsonResponse({
            'success': True,
            'message': f'Invitation {title_or_name} updated successfully'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@require_http_methods(["GET", "POST"])
def invitation_register_view(request, invitation_key):
    try:
        invitation = Invitation.objects.get(invitation_key=invitation_key, status='active')
    except Invitation.DoesNotExist:
        return render(request, 'invitations/invitation_invalid.html', status=404)

    if invitation.expiry_date and invitation.expiry_date < date.today():
        return render(request, 'invitations/invitation_expired.html', {'invitation': invitation})

    if request.method == 'POST':
        # Example: Register the user (you can expand this logic as needed)
        invitation.registered_count = (invitation.registered_count or 0) + 1
        invitation.save()
        return render(request, 'invitations/invitation_registered.html', {'invitation': invitation})

    return render(request, 'invitations/invitation_register.html', {'invitation': invitation})


def invite_landing(request, invitation_key):
    # Validate invitation key
    invitation = get_object_or_404(Invitation, invitation_key=invitation_key)
    
    # Check if invitation is active and not expired
    if invitation.status != 'active' or invitation.expiry_date.date() < timezone.now().date():
        return render(request, 'invitations/invite_landing.html', {
            'error': 'This invitation is either inactive or has expired.',
            'invitation_key': invitation_key,
            'event_name': invitation.event.name if invitation.event else 'GITEX GLOBAL 2025'
        })

    # Check link limit
    if invitation.registered_count >= invitation.link_limit:
        return render(request, 'invitations/invite_landing.html', {
            'error': 'This invitation has reached its registration limit.',
            'invitation_key': invitation_key,
            'event_name': invitation.event.name if invitation.event else 'GITEX GLOBAL 2025'
        })

    # Handle register button click (POST request from landing page)
    if request.method == 'POST':
        return redirect('invitations:invite_register', invitation_key=invitation_key)

    return render(request, 'invitations/invite_landing.html', {
        'invitation_key': invitation_key,
        'event_name': invitation.event.name if invitation.event else 'GITEX GLOBAL 2025',
        'title_or_name': invitation.title_or_name,
        'personal_message': invitation.personal_message
    })


@csrf_exempt
@require_http_methods(["GET", "POST"])
def invite_register(request, invitation_key):
    if request.method == 'GET':
        # Validate invitation key
        invitation = get_object_or_404(Invitation, invitation_key=invitation_key)
        
        # Check if invitation is active and not expired
        if invitation.status != 'active' or invitation.expiry_date.date() < timezone.now().date():
            return render(request, 'invite_register.html', {
                'error': 'This invitation is either inactive or has expired.',
                'invitation_key': invitation_key
            })

        # Check link limit
        if invitation.registered_count >= invitation.link_limit:
            return render(request, 'invite_register.html', {
                'error': 'This invitation has reached its registration limit.',
                'invitation_key': invitation_key
            })

        phone = ''
        registered_user = invitation.registered_users.first()  # Get the first related RegisteredUser, if any
        if registered_user:
            phone = registered_user.phone

        return render(request, 'invitations/invite_register.html', {
            'invitation_key': invitation_key,
            'event_name': invitation.event.name if invitation.event else 'GITEX GLOBAL 2025',
            'full_name': invitation.title_or_name,
            'email': invitation.email,
            'phone': phone,
        })

    elif request.method == 'POST':
        # Validate invitation
        invitation = get_object_or_404(Invitation, invitation_key=invitation_key)
        
        if invitation.status != 'active':
            return JsonResponse({'error': 'This invitation is not active.'}, status=400)
        
        if invitation.expiry_date.date() < timezone.now().date():
            return JsonResponse({'error': 'This invitation has expired.'}, status=400)
        
        if invitation.registered_count >= invitation.link_limit:
            return JsonResponse({'error': 'This invitation has reached its registration limit.'}, status=400)

        # Extract form data
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()

        # Validate inputs
        if not full_name:
            return JsonResponse({'error': 'Full name is required.'}, status=400)
        
        if len(full_name) < 2:
            return JsonResponse({'error': 'Full name must be at least 2 characters long.'}, status=400)
        
        if len(full_name) > 100:
            return JsonResponse({'error': 'Full name cannot exceed 100 characters.'}, status=400)
        
        # Allow letters, spaces, hyphens, and apostrophes; require at least two words
        name_regex = r"^[a-zA-Z\s'\-]+$"
        if not re.match(name_regex, full_name):
            return JsonResponse({'error': 'Full name can only contain letters, spaces, hyphens, and apostrophes.'}, status=400)
        
        if len(full_name.split()) < 2:
            return JsonResponse({'error': 'Full name must include at least a first and last name.'}, status=400)
        
        # Ensure email matches invitation's email
        if email != invitation.email:
            return JsonResponse({'error': 'Email cannot be changed.'}, status=400)
        
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not email or not re.match(email_regex, email):
            return JsonResponse({'error': 'A valid email address is required.'}, status=400)
        
        phone_regex = r'^\+?\d{8,15}$'
        if not phone or not re.match(phone_regex, phone):
            return JsonResponse({'error': 'A valid phone number is required.'}, status=400)

        # Check for duplicate email
        if RegisteredUser.objects.filter(email=email).exists():
            return JsonResponse({'error': 'This email is already registered.'}, status=400)

        try:
            # Create RegisteredUser
            registered_user = RegisteredUser.objects.create(
                invitation=invitation,
                full_name=full_name,
                email=email,
                phone=phone
            )
            

            # Increment registered_count
            invitation.registered_count += 1
            invitation.status = 'expired'
            invitation.save()

            return JsonResponse({
                'message': 'Registration successful! Your visitor pass has been registered.',
                'full_name': full_name
            })

        except ValidationError as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    


from django.http import HttpResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from .tasks import export_invitations_task
from celery.result import AsyncResult
from django.http import JsonResponse



@csrf_exempt
@require_http_methods(["POST"])
def export_jobs(request):
    # Get filter parameters
    export_format = request.POST.get('format', 'csv')
    print('-------export_format------', request.POST, export_format )
    task = export_invitations_task.delay(export_format)
    return JsonResponse({'task_id': task.id})


class ExportInvitationsView(LoginRequiredMixin, View):
    def post(self, request):
        # Get filter parameters
        export_format = request.POST.get('format', 'csv')
        task = export_invitations_task(export_format)
        return JsonResponse({'task_id': task.id})
        
    #     # Get filtered invitations
    #     invitations = Invitation.objects.filter()
        
    #     # Apply filters
    #     # if keyword:
    #     #     invitations = invitations.filter(
    #     #         Q(guest_name__icontains=keyword) | 
    #     #         Q(guest_email__icontains=keyword) |
    #     #         Q(link_title__icontains=keyword)
    #     #     )
    #     if status:
    #         invitations = invitations.filter(status=status)
    #     if inv_type:
    #         if inv_type == 'link':
    #             invitations = invitations.filter(invitation_type='link')
    #         elif inv_type == 'personal':
    #             invitations = invitations.filter(invitation_type='personal')
    #     if date_filter:
    #         invitations = invitations.filter(expiry_date=date_filter)
        
    #     # Generate file based on format
    #     if export_format == 'csv':
    #         return self.export_csv(invitations)
    #     elif export_format == 'excel':
    #         return self.export_excel(invitations)
    #     elif export_format == 'pdf':
    #         return self.export_pdf(invitations)
    #     else:
    #         return HttpResponse('Invalid format', status=400)
    
    # def export_csv(self, invitations):
    #     response = HttpResponse(content_type='text/csv')
    #     response['Content-Disposition'] = f'attachment; filename="invitations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
    #     writer = csv.writer(response)
    #     writer.writerow(['Name', 'Email', 'Type', 'Ticket Type', 'Expiry Date', 'Limit', 'Registered', 'Status', 'Company', 'Invitation Link'])
        
    #     for inv in invitations:
    #         writer.writerow([
    #             inv.title_or_name ,
    #             inv.email or 'N/A',
    #             'Link' if inv.invite_type == 'link' else 'Personal',
    #             inv.ticket_class.ticket_type,
    #             inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
    #             inv.link_limit,
    #             inv.registered_count,
    #             inv.status.capitalize(),
    #             inv.company_name or 'N/A',
    #             # f"{request.scheme}://{request.get_host()}/invite/{inv.invitation_key}"
    #         ])
        
    #     return response
    
    # def export_excel(self, invitations):
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.title = "Invitations"
        
    #     # Header style
    #     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    #     header_font = Font(bold=True, color="FFFFFF")
        
    #     # Headers
    #     headers = ['Name', 'Email', 'Type', 'Ticket Type', 'Expiry Date', 'Limit', 'Registered', 'Status', 'Company', 'Invitation Link']
    #     ws.append(headers)
        
    #     # Style header row
    #     for cell in ws[1]:
    #         cell.fill = header_fill
    #         cell.font = header_font
    #         cell.alignment = Alignment(horizontal='center', vertical='center')
        
    #     # Data rows
    #     for inv in invitations:
    #         ws.append([
    #             inv.title_or_name,
    #             inv.email or 'N/A',
    #             'Link' if inv.invite_type == 'link' else 'Personal',
    #             inv.ticket_class.ticket_type,
    #             inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
    #             inv.link_limit,
    #             inv.registered_count,
    #             inv.status.capitalize(),
    #             inv.company_name or 'N/A',
    #             # f"{request.scheme}://{request.get_host()}/invite/{inv.invitation_key}"
    #         ])
        
    #     # Adjust column widths
    #     for column in ws.columns:
    #         max_length = 0
    #         column_letter = column[0].column_letter
    #         for cell in column:
    #             try:
    #                 if len(str(cell.value)) > max_length:
    #                     max_length = len(cell.value)
    #             except:
    #                 pass
    #         adjusted_width = min(max_length + 2, 50)
    #         ws.column_dimensions[column_letter].width = adjusted_width
        
    #     # Save to response
    #     response = HttpResponse(
    #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     )
    #     response['Content-Disposition'] = f'attachment; filename="invitations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    #     wb.save(response)
        
    #     return response
    
    # def export_pdf(self, invitations):
    #     buffer = io.BytesIO()
    #     doc = SimpleDocTemplate(buffer, pagesize=A4)
    #     elements = []
        
    #     # Title
    #     styles = getSampleStyleSheet()
    #     title = Paragraph(f"<b>Invitations Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}</b>", styles['Title'])
    #     elements.append(title)
        
    #     # Table data
    #     data = [['Name', 'Email', 'Type', 'Status', 'Expiry', 'Registered']]
        
    #     for inv in invitations:
    #         data.append([
    #             inv.title_or_name ,
    #             inv.email or 'N/A',
    #             'Link' if inv.invite_type == 'link' else 'Personal',
    #             inv.ticket_class.ticket_type,
    #             inv.status.capitalize(),
    #             inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
    #             str(inv.registered_count)
    #         ])
        
    #     # Create table
    #     table = Table(data)
    #     table.setStyle(TableStyle([
    #         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    #         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    #         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    #         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    #         ('FONTSIZE', (0, 0), (-1, 0), 10),
    #         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    #         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    #         ('GRID', (0, 0), (-1, -1), 1, colors.black)
    #     ]))
        
    #     elements.append(table)
    #     doc.build(elements)
        
    #     buffer.seek(0)
    #     response = HttpResponse(buffer, content_type='application/pdf')
    #     response['Content-Disposition'] = f'attachment; filename="invitations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
    #     return response
    

@csrf_exempt
@require_http_methods(["POST"])
def bulk_delete(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)

      
        item_ids = data.get('SelectedIds', [])
        print('-------------data--------', data)
        print('----------------item_ids--------', item_ids)
        Invitation.objects.filter(id__in=item_ids).delete()


        return JsonResponse({
                'message': 'Deletion successful.',
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def bulk_activate(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)

        print('-------------data--------', data)
        item_ids = data.get('SelectedIds', [])
        Invitation.objects.filter(id__in=item_ids).update(status='active')

        return JsonResponse({
                'message': 'Successfuly activated',
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def bulk_deactivate(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)

        print('-------------data--------', data)
        item_ids = data.get('SelectedIds', [])
        Invitation.objects.filter(id__in=item_ids).update(status='expired')

        return JsonResponse({
                'message': 'Deactivate successful.',
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    


@csrf_exempt
@require_http_methods(["POST"])
def bulk_send_invites(request):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)

        print('-------------data--------', data)
        item_ids = data.get('SelectedIds', [])
        invites = Invitation.objects.filter(id__in=item_ids)

        for invite in invites:
            send_invitation_email.delay(invite.id)

        return JsonResponse({
                'message': 'Send successful.',
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@csrf_exempt
@require_http_methods(["POST"])
def send_broadcast(request, pk):
    """
    Handle sending a single personalized invitation via AJAX POST.
    Generates multiple unique invitation links based on link_count.
    """
    try:
        data = json.loads(request.body)
        email = data.get('email', '')

        invitation = get_object_or_404(Invitation, pk=pk)
        invitation.email = email
        invitation.personal_message = data.get('personal_message', invitation.personal_message)
        invitation.save()

        send_invitation_email.delay(pk)

        return JsonResponse({
                'success': True,
                'message': 'Send successful.',
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

from celery.result import AsyncResult
from .tasks import export_invitations_task


@csrf_exempt
@require_http_methods(["POST"])
def export_invitations(request):
    try:
        data = json.loads(request.body)
        format = data.get('format')
        keyword = data.get('keyword')
        status = data.get('status')
        type = data.get('type')
        date = data.get('date')

        if format not in ['csv', 'excel', 'pdf']:
            return JsonResponse({'success': False, 'error': 'Invalid format'}, status=400)

        # Start Celery task
        task = export_invitations_task.delay(format, keyword, status, type, date)
        return JsonResponse({'success': True, 'task_id': task.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@require_http_methods(["GET"])
def check_export_status(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'status': 'FAILURE', 'error': 'No task ID provided'}, status=400)

    task = AsyncResult(task_id)
    if task.state == 'SUCCESS':
        result = task.get()
        return JsonResponse(result)
    elif task.state == 'FAILURE':
        return JsonResponse({'status': 'FAILURE', 'error': str(task.get(propagate=False))})
    else:
        return JsonResponse({'status': task.state})