from celery import shared_task
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from .models import Invitation, ExportJob



def export_csv(invitations, export_job, total_invitations):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Email', 'Type', 'Ticket Type', 'Expiry Date', 'Limit', 'Registered', 'Status', 'Company'])
    
    processed = 0
    for inv in invitations:
        writer.writerow([
            inv.title_or_name,
            inv.email or 'N/A',
            'Link' if inv.invite_type == 'link' else 'Personal',
            inv.ticket_class.ticket_type,
            inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
            inv.link_limit,
            inv.registered_count,
            inv.status.capitalize(),
            inv.company_name or 'N/A'
        ])
        processed += 1
        # Update progress (20-80% for data processing)
        if total_invitations > 0:
            export_job.progress = 20 + (processed / total_invitations * 60)
            export_job.save()
    
    return io.BytesIO(output.getvalue().encode('utf-8'))

def export_excel(invitations, export_job, total_invitations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invitations"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Name', 'Email', 'Type', 'Ticket Type', 'Expiry Date', 'Limit', 'Registered', 'Status', 'Company']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Update progress after headers
    export_job.progress = 30
    export_job.save()
    
    # Data rows
    processed = 0
    for inv in invitations:
        ws.append([
            inv.title_or_name,
            inv.email or 'N/A',
            'Link' if inv.invite_type == 'link' else 'Personal',
            inv.ticket_class.ticket_type,
            inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
            inv.link_limit,
            inv.registered_count,
            inv.status.capitalize(),
            inv.company_name or 'N/A'
        ])
        processed += 1
        # Update progress (30-80% for data processing)
        if total_invitations > 0:
            export_job.progress = 30 + (processed / total_invitations * 50)
            export_job.save()
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_pdf(invitations, export_job, total_invitations):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Invitations Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}</b>", styles['Title'])
    elements.append(title)
    
    # Update progress after title
    export_job.progress = 30
    export_job.save()
    
    # Table data
    data = [['Name', 'Email', 'Type', 'Ticket Type', 'Status', 'Expiry', 'Registered']]
    
    processed = 0
    for inv in invitations:
        data.append([
            inv.title_or_name,
            inv.email or 'N/A',
            'Link' if inv.invite_type == 'link' else 'Personal',
            inv.ticket_class.ticket_type,
            inv.status.capitalize(),
            inv.expiry_date.strftime('%Y-%m-%d') if inv.expiry_date else 'N/A',
            str(inv.registered_count)
        ])
        processed += 1
        # Update progress (30-80% for data processing)
        if total_invitations > 0:
            export_job.progress = 30 + (processed / total_invitations * 50)
            export_job.save()
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer